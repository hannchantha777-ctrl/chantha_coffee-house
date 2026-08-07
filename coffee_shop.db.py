# -*- coding: utf-8 -*-
"""
Migrate coffee_shop.xlsx → coffee_shop.db (SQLite)
Run this script ONCE only.
"""

import sqlite3
from pathlib import Path
from openpyxl import load_workbook

# ==================== PATH ====================
APP_DIR = Path(__file__).resolve().parent
EXCEL_PATH = APP_DIR / "coffee_shop.xlsx"          # ឬ APP_DIR / "data" / "coffee_shop.xlsx"
DB_PATH = APP_DIR / "coffee_shop.db"

def migrate():
    if not EXCEL_PATH.exists():
        print(f"❌ រកមិនឃើញ Excel file: {EXCEL_PATH}")
        print("   សូមដាក់ coffee_shop.xlsx ក្បែរ script នេះ ឬកែ path")
        return

    print(f"📂 កំពុងអាន: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, data_only=True)

    # ==================== បង្កើត Database ====================
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # លុបតារាងចាស់ (ប្រសិនបើមាន) ដើម្បីចាប់ផ្តើមថ្មី
    cur.execute("DROP TABLE IF EXISTS products")
    cur.execute("DROP TABLE IF EXISTS orders")

    # បង្កើតតារាង products
    cur.execute("""
        CREATE TABLE products (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            name_kh TEXT,
            category TEXT,
            price REAL,
            stock INTEGER,
            unit TEXT,
            active TEXT DEFAULT 'Yes'
        )
    """)

    # បង្កើតតារាង orders
    cur.execute("""
        CREATE TABLE orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER,
            date TEXT,
            time TEXT,
            product_id INTEGER,
            product_name TEXT,
            qty INTEGER,
            unit_price REAL,
            total REAL,
            payment TEXT,
            staff TEXT
        )
    """)

    # ==================== Migrate Products ====================
    if "Products" not in wb.sheetnames:
        print("❌ រកមិនឃើញ sheet 'Products'")
    else:
        ws = wb["Products"]
        products = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            try:
                pid = int(row[0])
                name = str(row[1] or "").strip()
                name_kh = str(row[2] or "").strip()
                category = str(row[3] or "Other").strip()
                price = float(row[4] or 0)
                stock = int(row[5] or 0)
                unit = str(row[6] or "cup").strip()
                active = str(row[7] or "Yes").strip()
                if active.lower() not in ("yes", "no"):
                    active = "Yes"

                products.append((pid, name, name_kh, category, price, stock, unit, active))
            except Exception as e:
                print(f"⚠️ រំលង product row: {row} → {e}")

        cur.executemany("""
            INSERT INTO products (id, name, name_kh, category, price, stock, unit, active)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, products)
        print(f"✅ Products: បានផ្ទេរ {len(products)} មុខ")

    # ==================== Migrate Orders ====================
    if "Orders" not in wb.sheetnames:
        print("❌ រកមិនឃើញ sheet 'Orders'")
    else:
        ws = wb["Orders"]
        orders = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            try:
                order_id = int(row[0])
                date_ = str(row[1] or "")
                time_ = str(row[2] or "")
                product_id = int(row[3] or 0)
                product_name = str(row[4] or "")
                qty = int(row[5] or 0)
                unit_price = float(row[6] or 0)
                total = float(row[7] or 0)
                payment = str(row[8] or "")
                staff = str(row[9] or "")

                orders.append((
                    order_id, date_, time_, product_id, product_name,
                    qty, unit_price, total, payment, staff
                ))
            except Exception as e:
                print(f"⚠️ រំលង order row: {row} → {e}")

        cur.executemany("""
            INSERT INTO orders (order_id, date, time, product_id, product_name, qty, unit_price, total, payment, staff)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, orders)
        print(f"✅ Orders: បានផ្ទេរ {len(orders)} ជួរ")

    conn.commit()
    conn.close()
    wb.close()

    print("\n🎉 បំលែងបានជោគជ័យ!")
    print(f"📁 Database ថ្មី: {DB_PATH}")
    print("ឥឡូវអ្នកអាចរត់ app.py បានហើយ។")

if __name__ == "__main__":
    migrate()