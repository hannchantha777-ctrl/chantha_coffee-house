# -*- coding: utf-8 -*-
"""
Coffee House POS — Streamlit Version (Full)
Beautiful Modern UI + Excel Backend
Supports: Add / Edit / Soft Delete Products
+ Multi QR Code (ABA, Wing, ACLEDA, Canadia)
+ Receipt Popup
"""

import os
import shutil
import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook
import io

try:
    import qrcode
    from PIL import Image
    HAS_QR = True
except ImportError:
    HAS_QR = False


# ==================== PATH ====================
def _app_dir():
    return Path(__file__).resolve().parent

APP_DIR = _app_dir()
DATA_PATH = APP_DIR / "data" / "coffee_shop.xlsx"
DATA_DIR = APP_DIR / "data"

# Mapping payment → QR filename
QR_FILES = {
    "ABA": "aba_qr.png",
    "Wing": "wing_qr.png",
    "ACLEDA": "acleda_qr.png",
    "Canadia": "canadia_qr.png",
    "Other": "other_qr.png",
}


# ==================== PAGE CONFIG ====================
st.set_page_config(
    page_title="Coffee House POS",
    page_icon="☕",
    layout="wide",
    initial_sidebar_state="expanded"
)


# ==================== CUSTOM CSS ====================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+Khmer:wght@400;500;600;700&family=Poppins:wght@400;500;600;700&display=swap');
    
    html, body, [class*="css"], .stApp {
        font-family: 'Poppins', 'Noto Sans Khmer', sans-serif !important;
    }
    
    .stApp {
        background: linear-gradient(135deg, #faf6f1 0%, #f5efe6 100%);
    }
    
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #4a3728 0%, #3a2a1f 100%) !important;
        min-width: 280px !important;
    }
    
    [data-testid="stSidebar"] p, 
    [data-testid="stSidebar"] span, 
    [data-testid="stSidebar"] label, 
    [data-testid="stSidebar"] h1, 
    [data-testid="stSidebar"] h2, 
    [data-testid="stSidebar"] h3 {
        color: #f5efe6 !important;
    }
    
    /* Large Menu */
    [data-testid="stSidebar"] .stRadio label {
        font-size: 1.2rem !important;
        padding: 12px 10px !important;
        margin: 4px 0 !important;
        border-radius: 10px !important;
        transition: background 0.2s;
        display: block;
    }
    [data-testid="stSidebar"] .stRadio label:hover {
        background: rgba(255,255,255,0.12) !important;
    }
    
    .sidebar-logo {
        text-align: center;
        padding: 10px 0 6px 0;
    }
    .sidebar-logo .logo-icon {
        font-size: 3rem;
        line-height: 1;
        display: block;
    }
    .sidebar-logo .logo-title {
        font-size: 1.4rem;
        font-weight: 700;
        color: #fff !important;
        margin-top: 4px;
    }
    .sidebar-logo .logo-sub {
        font-size: 0.85rem;
        opacity: 0.85;
        color: #f5efe6 !important;
    }
    
    div[data-testid="stMetric"] {
        background: #ffffff !important;
        border: 1px solid #e8dfd4;
        border-radius: 16px;
        padding: 18px 20px;
        box-shadow: 0 4px 15px rgba(74, 55, 40, 0.05);
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    
    div[data-testid="stMetric"]:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(74, 55, 40, 0.1);
    }
    
    div[data-testid="stMetric"] label,
    div[data-testid="stMetric"] [data-testid="stMetricLabel"] p {
        color: #6F4E37 !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
    }
    
    div[data-testid="stMetric"] [data-testid="stMetricValue"] div {
        color: #4a3728 !important;
        font-size: 1.8rem !important;
        font-weight: 700 !important;
    }
    
    h1, h2, h3, h4, h5, h6 {
        color: #4a3728 !important;
        font-weight: 700;
    }
    
    p, label, span {
        color: #33271e;
    }
    
    .stTextInput label, .stSelectbox label, .stNumberInput label, .stTextArea label {
        color: #4a3728 !important;
        font-weight: 600 !important;
    }

    .stButton > button {
        background: linear-gradient(135deg, #6F4E37 0%, #5a3e2b 100%) !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 12px !important;
        padding: 0.6rem 1.4rem !important;
        font-weight: 600 !important;
        transition: all 0.25s ease !important;
        box-shadow: 0 4px 12px rgba(111, 78, 55, 0.25) !important;
    }
    
    .stButton > button:hover {
        background: linear-gradient(135deg, #5a3e2b 0%, #4a3728 100%) !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 16px rgba(111, 78, 55, 0.35) !important;
        color: #ffffff !important;
    }

    .stButton > button p {
        color: #ffffff !important;
    }
    
    .product-card {
        background: #ffffff;
        border-radius: 16px;
        padding: 16px;
        border: 1px solid #e8dfd4;
        box-shadow: 0 4px 12px rgba(74, 55, 40, 0.06);
        margin-bottom: 12px;
        transition: all 0.2s ease;
    }
    
    .product-card:hover {
        box-shadow: 0 8px 20px rgba(74, 55, 40, 0.12);
        transform: translateY(-2px);
    }
    
    .low-stock {
        background: #fff8f0 !important;
        border-left: 5px solid #d35400 !important;
        padding: 12px 16px;
        border-radius: 8px;
        margin: 8px 0;
        box-shadow: 0 2px 8px rgba(211, 84, 0, 0.08);
        color: #4a3728 !important;
        font-weight: 500;
    }

    .low-stock span, .low-stock strong {
        color: #d35400 !important;
    }
    
    .cart-box {
        background: #ffffff;
        border-radius: 16px;
        padding: 20px;
        border: 1px solid #e8dfd4;
        box-shadow: 0 4px 16px rgba(74, 55, 40, 0.08);
        position: sticky;
        top: 80px;
    }
    
    /* Receipt Popup Style */
    .receipt-overlay {
        background: white;
        border: 2px dashed #6F4E37;
        border-radius: 16px;
        padding: 24px 28px;
        max-width: 460px;
        margin: 0 auto 20px auto;
        box-shadow: 0 12px 40px rgba(74, 55, 40, 0.22);
        font-family: 'Courier New', monospace;
        animation: popIn 0.35s ease-out;
    }
    @keyframes popIn {
        from { opacity: 0; transform: scale(0.92) translateY(-18px); }
        to   { opacity: 1; transform: scale(1) translateY(0); }
    }
    .receipt-title {
        text-align: center;
        font-weight: 700;
        font-size: 1.1rem;
        letter-spacing: 1px;
        color: #4a3728;
        margin-bottom: 2px;
    }
    .receipt-sub {
        text-align: center;
        font-size: 0.92rem;
        color: #6F4E37;
        margin-bottom: 10px;
    }
    .receipt-line {
        border-top: 1px dashed #c4a882;
        margin: 8px 0;
    }
    .receipt-row {
        display: flex;
        justify-content: space-between;
        font-size: 0.92rem;
        margin: 2px 0;
        color: #333;
    }
    .receipt-total {
        font-weight: 700;
        font-size: 1.15rem;
        text-align: center;
        margin: 10px 0 4px 0;
        color: #4a3728;
    }
    .receipt-thanks {
        text-align: center;
        font-size: 0.92rem;
        margin-top: 6px;
        color: #6F4E37;
    }
</style>
""", unsafe_allow_html=True)


# ==================== DATA FUNCTIONS ====================
def load_products(include_inactive=False):
    if not DATA_PATH.exists():
        return []
    wb = load_workbook(DATA_PATH, data_only=True)
    ws = wb["Products"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    products = []
    for r in rows:
        if r[0] is None:
            continue
        active = str(r[7]).strip() if r[7] else "Yes"
        if not include_inactive and active != "Yes":
            continue
        products.append({
            "id": int(r[0]),
            "name": str(r[1]),
            "name_kh": str(r[2] or ""),
            "category": str(r[3] or ""),
            "price": float(r[4] or 0),
            "stock": int(r[5] or 0),
            "unit": str(r[6] or "cup"),
            "active": active,
        })
    wb.close()
    return products


def load_orders():
    if not DATA_PATH.exists():
        return []
    wb = load_workbook(DATA_PATH, data_only=True)
    ws = wb["Orders"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    orders = []
    for r in rows:
        if r[0] is None:
            continue
        orders.append({
            "order_id": int(r[0]),
            "date": str(r[1]),
            "time": str(r[2]),
            "product_id": int(r[3]) if r[3] else 0,
            "product_name": str(r[4] or ""),
            "qty": int(r[5] or 0),
            "unit_price": float(r[6] or 0),
            "total": float(r[7] or 0),
            "payment": str(r[8] or ""),
            "staff": str(r[9] or ""),
        })
    wb.close()
    return orders


def get_next_order_id():
    orders = load_orders()
    if not orders:
        return 1001
    return max(o["order_id"] for o in orders) + 1


def _safe_save(wb, path):
    path = Path(path)
    tmp = path.with_suffix(".tmp.xlsx")
    try:
        wb.save(str(tmp))
        wb.close()
        if path.exists():
            try:
                os.remove(str(path))
            except PermissionError:
                raise PermissionError(
                    "Permission denied\n\n"
                    "សូមបិទ Excel / WPS ដែលកំពុងបើក file:\n"
                    f"{path}\n\n"
                    "រួចសាកម្តងទៀត។"
                )
        shutil.move(str(tmp), str(path))
    except PermissionError:
        raise
    except Exception:
        try:
            wb.save(str(path))
            wb.close()
        except PermissionError:
            raise PermissionError(
                "Permission denied\n\n"
                "សូមបិទ Excel / WPS ដែលកំពុងបើក file:\n"
                f"{path}\n\n"
                "រួចសាកម្តងទៀត។"
            )
    finally:
        if tmp.exists():
            try:
                os.remove(str(tmp))
            except Exception:
                pass


def save_order(lines):
    wb = load_workbook(DATA_PATH)
    ws = wb["Orders"]
    for row in lines:
        ws.append(row)

    ws_p = wb["Products"]
    id_to_row = {}
    for i, row in enumerate(ws_p.iter_rows(min_row=2), start=2):
        if row[0].value is not None:
            id_to_row[int(row[0].value)] = i

    for line in lines:
        pid, qty = line[3], line[5]
        if pid in id_to_row:
            cell = ws_p.cell(id_to_row[pid], 6)
            current = int(cell.value or 0)
            cell.value = max(0, current - int(qty))

    _safe_save(wb, DATA_PATH)


def update_stock(product_id, new_stock):
    wb = load_workbook(DATA_PATH)
    ws = wb["Products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            row[5].value = int(new_stock)
            break
    _safe_save(wb, DATA_PATH)


def add_product(name, name_kh, category, price, stock, unit):
    wb = load_workbook(DATA_PATH)
    ws = wb["Products"]
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            max_id = max(max_id, int(row[0]))
    new_id = max_id + 1
    ws.append([new_id, name, name_kh, category, float(price), int(stock), unit, "Yes"])
    _safe_save(wb, DATA_PATH)
    return new_id


def update_product(product_id, name, name_kh, category, price, stock, unit):
    wb = load_workbook(DATA_PATH)
    ws = wb["Products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            row[1].value = name
            row[2].value = name_kh
            row[3].value = category
            row[4].value = float(price)
            row[5].value = int(stock)
            row[6].value = unit
            break
    _safe_save(wb, DATA_PATH)


def soft_delete_product(product_id):
    wb = load_workbook(DATA_PATH)
    ws = wb["Products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            row[7].value = "No"
            break
    _safe_save(wb, DATA_PATH)


def restore_product(product_id):
    wb = load_workbook(DATA_PATH)
    ws = wb["Products"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product_id:
            row[7].value = "Yes"
            break
    _safe_save(wb, DATA_PATH)


def generate_qr(data: str, size=200):
    if not HAS_QR:
        return None
    qr = qrcode.QRCode(version=1, box_size=6, border=2)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#1a1a6c", back_color="white")
    img = img.resize((size, size))
    return img


def get_qr_image(order_id, total, payment):
    """Prefer real bank QR file, else generate text QR."""
    filename = QR_FILES.get(payment)
    if filename:
        qr_path = DATA_DIR / filename
        if qr_path.exists():
            return str(qr_path)

    # Fallback generated QR
    data = f"COFFEE HOUSE\nOrder #{order_id}\nTotal: ${total:.2f}\nPayment: {payment}"
    return generate_qr(data)


def check_available_qrs():
    available = {}
    for pay, fname in QR_FILES.items():
        path = DATA_DIR / fname
        available[pay] = path.exists()
    return available


# ==================== SESSION STATE ====================
if "cart" not in st.session_state:
    st.session_state.cart = {}
if "last_receipt" not in st.session_state:
    st.session_state.last_receipt = None
if "edit_product_id" not in st.session_state:
    st.session_state.edit_product_id = None


# ==================== SIDEBAR ====================
with st.sidebar:
    st.markdown("""
    <div class="sidebar-logo">
        <span class="logo-icon">☕</span>
        <div class="logo-title">Coffee House</div>
        <div class="logo-sub">POS System</div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")
    
    menu = st.radio(
        "ជ្រើសរើសមុខងារ",
        [
            "🏠 ទំព័រដើម",
            "🛒 លក់ថ្មី",
            "📦 ផលិតផល",
            "📋 ប្រវត្តិលក់",
            "📊 របាយការណ៍",
            "⚙️ QR Code",
        ],
        label_visibility="collapsed"
    )
    
    st.markdown("---")
    st.caption("© Coffee House POS")


# ==================== PAGES ====================

# ---------- DASHBOARD ----------
if menu == "🏠 ទំព័រដើម":
    st.markdown("## 🏠 ទំព័រដើម")
    st.markdown("ទិដ្ឋភាពទូទៅនៃហាងថ្ងៃនេះ")
    
    products = load_products()
    orders = load_orders()
    today = date.today().strftime("%Y-%m-%d")
    today_orders = [o for o in orders if o["date"] == today]
    revenue = sum(o["total"] for o in today_orders)
    order_ids = set(o["order_id"] for o in today_orders)
    low = [p for p in products if p["stock"] <= 20]
    
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("💰 ចំណូលថ្ងៃនេះ", f"${revenue:.2f}")
    col2.metric("🧾 Order ថ្ងៃនេះ", len(order_ids))
    col3.metric("📦 ផលិតផល", len(products))
    col4.metric("⚠️ ស្តុកទាប", len(low))
    
    st.markdown("### ⚠️ ផលិតផលស្តុកទាប")
    
    if low:
        for p in low:
            st.markdown(
                f"""
                <div class="low-stock">
                    <strong>{p['name']}</strong> ({p['name_kh']}) — 
                    ស្តុកនៅសល់: <b style="color:#e67e22;">{p['stock']} {p['unit']}</b>
                </div>
                """,
                unsafe_allow_html=True
            )
    else:
        st.success("ស្តុកគ្រប់គ្រាន់ ✅")


# ---------- ORDER ----------
elif menu == "🛒 លក់ថ្មី":
    st.markdown("## 🛒 លក់ថ្មី")
    st.markdown("ជ្រើសរើសផលិតផល និងបញ្ជាក់ការលក់")
    
        # ===== Receipt Popup (បង្ហាញនៅខាងលើ) =====
    if st.session_state.last_receipt:
        rec = st.session_state.last_receipt

        # បង្កើតបញ្ជីទំនិញជា text ស្អាត (មិនខូច HTML)
        items_lines = ""
        for it in rec["items"]:
            items_lines += f"{it['name']}\n"
            items_lines += f"  {it['qty']} x ${it['price']:.2f}  =  ${it['subtotal']:.2f}\n"

        receipt_text = f"""====================================
     COFFEE HOUSE កាហ្វេហោស
        *** RECEIPT / វិក័យប័ត្រ ***
====================================
Order # : {rec['order_id']}
Date    : {rec['date_time']}
Staff   : {rec['staff']}
Payment : {rec['payment']}
------------------------------------
{items_lines}------------------------------------
TOTAL              ${rec['total']:.2f}
====================================
     សូមអរគុណ!  Thank you!
===================================="""

        col_rec, col_qr = st.columns([1.4, 1])

        with col_rec:
            st.markdown("### 🧾 វិក័យប័ត្រ")
            st.markdown(
                f"""
                <div class="receipt-overlay">
                    <pre style="margin:0; font-family:'Courier New',monospace; font-size:0.95rem; white-space:pre-wrap; color:#2b1d14;">{receipt_text}</pre>
                </div>
                """,
                unsafe_allow_html=True
            )

        with col_qr:
            st.markdown(f"### QR Code — {rec['payment']}")
            qr_img = get_qr_image(rec["order_id"], rec["total"], rec["payment"])
            if qr_img is not None:
                st.image(qr_img, width=230)
            else:
                st.warning("មិនអាចបង្ហាញ QR")

            st.caption(f"Order #{rec['order_id']} • ${rec['total']:.2f} • {rec['payment']}")

            avail = check_available_qrs()
            if rec["payment"] in avail and avail[rec["payment"]]:
                st.success(f"✅ ប្រើ QR {rec['payment']} ពិត")
            else:
                st.info(f"ℹ️ ប្រើ QR ស្វ័យប្រវត្តិ (ដាក់ {QR_FILES.get(rec['payment'], '?')} ដើម្បីប្រើ QR ពិត)")

            # ប៊ូតុង Print + បិទ
            col_p1, col_p2 = st.columns(2)
            with col_p1:
                # ប៊ូតុង Print (ប្រើ browser print)
                st.markdown(
                    f"""
                    <a href="javascript:window.print()" style="
                        display:inline-block;
                        width:100%;
                        text-align:center;
                        background:linear-gradient(135deg,#6F4E37,#5a3e2b);
                        color:white;
                        padding:0.6rem 1rem;
                        border-radius:12px;
                        text-decoration:none;
                        font-weight:600;
                        box-shadow:0 4px 12px rgba(111,78,55,0.25);
                    ">🖨️ Print</a>
                    """,
                    unsafe_allow_html=True
                )
            with col_p2:
                if st.button("បិទ / បន្តលក់", use_container_width=True, type="primary"):
                    st.session_state.last_receipt = None
                    st.rerun()

            # ប៊ូតុងទាញយកវិក័យប័ត្រជា .txt
            st.download_button(
                label="💾 ទាញយកវិក័យប័ត្រ (.txt)",
                data=receipt_text,
                file_name=f"receipt_{rec['order_id']}.txt",
                mime="text/plain",
                use_container_width=True
            )

        st.markdown("---")
        st.success(f"✅ លក់ជោគជ័យ! Order ID: **{rec['order_id']}** | សរុប: **${rec['total']:.2f}**")
        st.markdown("---")
    
    products = load_products()
    
    left, right = st.columns([1.6, 1])
    
    with left:
        st.subheader("ជ្រើសរើសផលិតផល")
        
        categories = ["ទាំងអស់"] + sorted(list(set(p["category"] for p in products)))
        selected_cat = st.selectbox("ប្រភេទ", categories, key="cat_filter")
        
        filtered = products if selected_cat == "ទាំងអស់" else [p for p in products if p["category"] == selected_cat]
        
        cols = st.columns(3)
        for idx, p in enumerate(filtered):
            with cols[idx % 3]:
                with st.container():
                    st.markdown(
                        f"""
                        <div class="product-card">
                            <div style="font-weight:600; font-size:1.05rem; color:#4a3728;">{p['name']}</div>
                            <div style="color:#8b7355; font-size:0.9rem;">{p['name_kh']}</div>
                            <div style="margin-top:8px; font-size:1.15rem; font-weight:700; color:#6F4E37;">${p['price']:.2f}</div>
                            <div style="color:#888; font-size:0.85rem;">ស្តុក: {p['stock']} {p['unit']}</div>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
                    if st.button("➕ បន្ថែម", key=f"add_{p['id']}", use_container_width=True):
                        if p["stock"] <= 0:
                            st.warning(f"{p['name']} អស់ស្តុកហើយ")
                        else:
                            current_qty = st.session_state.cart.get(p["id"], {}).get("qty", 0)
                            if current_qty >= p["stock"]:
                                st.warning(f"ស្តុកនៅសល់តែ {p['stock']}")
                            else:
                                if p["id"] in st.session_state.cart:
                                    st.session_state.cart[p["id"]]["qty"] += 1
                                else:
                                    st.session_state.cart[p["id"]] = {
                                        "name": p["name"],
                                        "price": p["price"],
                                        "qty": 1
                                    }
                                st.rerun()
    
    with right:
        st.markdown('<div class="cart-box">', unsafe_allow_html=True)
        st.subheader("🛒 កន្ត្រក់")
        
        if not st.session_state.cart:
            st.info("កន្ត្រក់ទទេ")
        else:
            total = 0
            for pid, item in st.session_state.cart.items():
                sub = item["price"] * item["qty"]
                total += sub
                c1, c2, c3 = st.columns([3, 1, 1])
                c1.write(f"**{item['name']}**")
                c2.write(f"x{item['qty']}")
                c3.write(f"${sub:.2f}")
            
            st.markdown(f"### សរុប: **${total:.2f}**")
            
            payment = st.selectbox(
                "វិធីបង់ប្រាក់",
                ["Cash", "ABA", "Wing", "ACLEDA", "Canadia", "Other"],
                help="QR នឹងប្តូរតាមវិធីបង់ប្រាក់ដែលអ្នកជ្រើស"
            )
            staff = st.selectbox("បុគ្គលិក", ["Sokha", "Dara", "Pisey", "Other"])
            
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("✅ បញ្ជាក់ការលក់", use_container_width=True, type="primary"):
                    now = datetime.now()
                    order_id = get_next_order_id()
                    lines = []
                    items_for_receipt = []
                    total = 0.0
                    for pid, item in st.session_state.cart.items():
                        sub = round(item["price"] * item["qty"], 2)
                        total += sub
                        lines.append([
                            order_id,
                            now.strftime("%Y-%m-%d"),
                            now.strftime("%H:%M"),
                            pid,
                            item["name"],
                            item["qty"],
                            item["price"],
                            sub,
                            payment,
                            staff,
                        ])
                        items_for_receipt.append({
                            "name": item["name"],
                            "qty": item["qty"],
                            "price": item["price"],
                            "subtotal": sub,
                        })
                    try:
                        save_order(lines)
                        st.session_state.last_receipt = {
                            "order_id": order_id,
                            "date_time": now.strftime("%Y-%m-%d %H:%M"),
                            "staff": staff,
                            "payment": payment,
                            "total": total,
                            "items": items_for_receipt,
                        }
                        st.session_state.cart = {}
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))
            
            with col_btn2:
                if st.button("🗑️ សម្អាត", use_container_width=True):
                    st.session_state.cart = {}
                    st.rerun()
        
        st.markdown('</div>', unsafe_allow_html=True)


# ---------- PRODUCTS ----------
elif menu == "📦 ផលិតផល":
    st.markdown("## 📦 គ្រប់គ្រងផលិតផល")
    st.markdown("មើល • កែប្រែ • លុប ផលិតផល")
    
    tab1, tab2, tab3 = st.tabs(["📋 បញ្ជីផលិតផល", "✏️ កែប្រែ / លុប", "➕ បន្ថែមថ្មី"])
    
    with tab1:
        products = load_products()
        if products:
            df = pd.DataFrame(products)
            df = df[["id", "name", "name_kh", "category", "price", "stock", "unit"]]
            df.columns = ["ID", "Name", "ខ្មែរ", "Category", "Price", "Stock", "Unit"]
            st.dataframe(
                df.style.format({"Price": "${:.2f}"}),
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("មិនមានផលិតផលទេ")
    
    with tab2:
        products = load_products()
        inactive = load_products(include_inactive=True)
        inactive = [p for p in inactive if p["active"] != "Yes"]
        
        st.subheader("✏️ កែប្រែផលិតផល")
        
        if not products:
            st.warning("មិនមានផលិតផលដើម្បីកែប្រែ")
        else:
            options = {f"{p['id']} - {p['name']} ({p['name_kh']})": p for p in products}
            selected_label = st.selectbox("ជ្រើសរើសផលិតផលដើម្បីកែប្រែ", list(options.keys()))
            selected = options[selected_label]
            
            with st.form("edit_product_form"):
                col_a, col_b = st.columns(2)
                with col_a:
                    name = st.text_input("Name (EN)*", value=selected["name"])
                    name_kh = st.text_input("ឈ្មោះខ្មែរ", value=selected["name_kh"])
                    category = st.text_input("Category", value=selected["category"])
                with col_b:
                    price = st.number_input("Price ($)", min_value=0.0, value=float(selected["price"]), step=0.25)
                    stock = st.number_input("Stock", min_value=0, value=int(selected["stock"]), step=1)
                    unit = st.text_input("Unit", value=selected["unit"])
                
                col_save, col_delete = st.columns(2)
                with col_save:
                    submitted = st.form_submit_button("💾 រក្សាទុកការកែប្រែ", use_container_width=True)
                with col_delete:
                    delete_clicked = st.form_submit_button("🗑️ លុបផលិតផល", use_container_width=True)
                
                if submitted:
                    if not name.strip():
                        st.error("ត្រូវការឈ្មោះ")
                    else:
                        try:
                            update_product(
                                selected["id"],
                                name.strip(),
                                name_kh.strip(),
                                category.strip() or "Other",
                                price,
                                stock,
                                unit.strip() or "cup"
                            )
                            st.success("បានកែប្រែផលិតផលជោគជ័យ!")
                            st.rerun()
                        except Exception as e:
                            st.error(str(e))
                
                if delete_clicked:
                    try:
                        soft_delete_product(selected["id"])
                        st.success(f"បានលុប «{selected['name']}» ជោគជ័យ (Soft Delete)")
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))
        
        if inactive:
            st.markdown("---")
            st.subheader("♻️ ផលិតផលដែលបានលុប (អាចស្តារបាន)")
            for p in inactive:
                col1, col2 = st.columns([4, 1])
                with col1:
                    st.write(f"**{p['name']}** ({p['name_kh']}) — ID: {p['id']}")
                with col2:
                    if st.button("ស្តារ", key=f"restore_{p['id']}"):
                        try:
                            restore_product(p["id"])
                            st.success(f"បានស្តារ «{p['name']}»")
                            st.rerun()
                        except Exception as e:
                            st.error(str(e))
    
    with tab3:
        st.subheader("➕ បន្ថែមផលិតផលថ្មី")
        with st.form("add_product_form"):
            col1, col2 = st.columns(2)
            with col1:
                name = st.text_input("Name (EN)*")
                name_kh = st.text_input("ឈ្មោះខ្មែរ")
                category = st.text_input("Category", value="Coffee")
            with col2:
                price = st.number_input("Price ($)", min_value=0.0, value=1.5, step=0.25)
                stock = st.number_input("Stock", min_value=0, value=50, step=1)
                unit = st.text_input("Unit", value="cup")
            
            if st.form_submit_button("រក្សាទុកផលិតផល", use_container_width=True):
                if not name.strip():
                    st.error("ត្រូវការឈ្មោះ")
                else:
                    try:
                        add_product(
                            name.strip(),
                            name_kh.strip(),
                            category.strip() or "Other",
                            price,
                            stock,
                            unit.strip() or "cup"
                        )
                        st.success("បានបន្ថែមផលិតផលជោគជ័យ!")
                        st.rerun()
                    except Exception as e:
                        st.error(str(e))


# ---------- HISTORY ----------
elif menu == "📋 ប្រវត្តិលក់":
    st.markdown("## 📋 ប្រវត្តិការលក់")
    st.markdown("ប្រវត្តិការលក់ទាំងអស់")
    
    orders = load_orders()
    if not orders:
        st.info("មិនទាន់មានទិន្នន័យ")
    else:
        orders = sorted(orders, key=lambda o: (o["date"], o["time"]), reverse=True)
        df = pd.DataFrame(orders)
        df = df[["order_id", "date", "time", "product_name", "qty", "total", "payment", "staff"]]
        df.columns = ["OrderID", "Date", "Time", "Product", "Qty", "Total", "Payment", "Staff"]
        
        st.dataframe(
            df.style.format({"Total": "${:.2f}"}),
            use_container_width=True,
            hide_index=True
        )
        
        total = sum(o["total"] for o in orders)
        st.markdown(f"**សរុបទាំងអស់: ${total:.2f}**  |  ចំនួនជួរ: {len(orders)}")


# ---------- REPORT ----------
elif menu == "📊 របាយការណ៍":
    st.markdown("## 📊 របាយការណ៍")
    st.markdown("វិភាគចំណូល និងផលិតផលលក់ដាច់")
    
    orders = load_orders()
    if not orders:
        st.info("មិនទាន់មានទិន្នន័យ")
    else:
        daily = {}
        for o in orders:
            daily[o["date"]] = daily.get(o["date"], 0) + o["total"]
        
        st.subheader("💰 ចំណូលប្រចាំថ្ងៃ")
        daily_df = pd.DataFrame([
            {"Date": d, "Revenue": v} for d, v in sorted(daily.items(), reverse=True)
        ])
        st.dataframe(
            daily_df.style.format({"Revenue": "${:.2f}"}),
            use_container_width=True,
            hide_index=True
        )
        
        st.bar_chart(daily_df.set_index("Date")["Revenue"])
        
        st.subheader("🏆 ផលិតផលលក់ដាច់")
        prod = {}
        for o in orders:
            if o["product_name"] not in prod:
                prod[o["product_name"]] = {"qty": 0, "rev": 0}
            prod[o["product_name"]]["qty"] += o["qty"]
            prod[o["product_name"]]["rev"] += o["total"]
        
        prod_df = pd.DataFrame([
            {"Product": name, "Qty": v["qty"], "Revenue": v["rev"]}
            for name, v in sorted(prod.items(), key=lambda x: -x[1]["qty"])
        ])
        st.dataframe(
            prod_df.style.format({"Revenue": "${:.2f}"}),
            use_container_width=True,
            hide_index=True
        )


# ---------- QR CODE MANAGEMENT ----------
elif menu == "⚙️ QR Code":
    st.markdown("## ⚙️ គ្រប់គ្រង QR Code បង់ប្រាក់")
    
    st.markdown("""
    ### របៀបដាក់ QR Code ពិត
    1. បើក App ធនាគារ (ABA / Wing / ACLEDA / Canadia)
    2. ចូល **Receive Money** ឬ **KHQR**
    3. Screenshot / Download QR
    4. ផ្លាស់ឈ្មោះ file ឱ្យត្រូវតាមតារាងខាងក្រោម
    5. ដាក់ file ក្នុង folder **`data/`**
    """)
    
    st.markdown("---")
    st.subheader("📋 តារាងឈ្មោះ File QR")
    
    avail = check_available_qrs()
    table_data = []
    for pay, fname in QR_FILES.items():
        status = "✅ មាន" if avail.get(pay) else "❌ មិនទាន់មាន"
        table_data.append({
            "វិធីបង់ប្រាក់": pay,
            "ឈ្មោះ File": fname,
            "ស្ថានភាព": status
        })
    
    st.dataframe(pd.DataFrame(table_data), use_container_width=True, hide_index=True)
    
    st.markdown("---")
    st.subheader("👀 មើល QR ដែលមានស្រាប់")
    
    cols = st.columns(3)
    idx = 0
    for pay, fname in QR_FILES.items():
        path = DATA_DIR / fname
        if path.exists():
            with cols[idx % 3]:
                st.markdown(f"**{pay}**")
                st.image(str(path), width=180)
                st.caption(fname)
            idx += 1
    
    if idx == 0:
        st.info("មិនទាន់មាន QR file ណាមួយទេ។ សូមដាក់ `aba_qr.png` ។ល។ ក្នុង folder `data/`")
    
    st.markdown("---")
    st.info("មិនទាន់មាន QR file ណាមួយទេ។ សូមដាក់ `wing_qr.png` ។ល។ ក្នុង folder `data/`") 
    st.info("មិនទាន់មាន QR file ណាមួយទេ។ សូមដាក់ `acleda_qr.png` ។ល។ ក្នុង folder `data/`")